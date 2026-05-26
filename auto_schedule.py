# -*- coding: utf-8 -*-
"""
自动给实验室排杂交 pool + 上机分 line。

输入:  `输入文件.xlsx`
输出:  `输出文件.xlsx`（覆盖写入；列结构与模板一致，并合并"杂交编号"列相同的连续单元格）

规则来源: `杂交和上机规则（new）.docx`
"""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
import sys
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

from hybrid_group_mapping import PREFIX_TO_GROUP


INPUT_COLUMNS = [
    "样本编号",
    "报告截止时间",
    "实验编号",
    "预文库浓度（ng/μL）",
    "预文库等级",
    "预文库取样体积（μL）",
    "孔位",
    "文库index编号",
    "数据量（G）",
]

OUTPUT_COLUMNS = INPUT_COLUMNS + ["杂交编号", "提示信息", "line号"]


def normalize_input_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """兼容线下导出表头：标本编号≈样本编号；缺少报告截止时间时用占位日期。"""
    out = df.copy()
    out.attrs = dict(getattr(df, "attrs", {}) or {})
    if "样本编号" not in out.columns and "标本编号" in out.columns:
        out["样本编号"] = out["标本编号"]
    if "报告截止时间" not in out.columns:
        out["报告截止时间"] = pd.Timestamp("2099-12-31 23:59:59")
        out.attrs["deadline_column_missing"] = True
        print(
            "[WARN] 输入表无「报告截止时间」列：杂交/排机顺序按实验编号含「#」优先；"
            "已写入占位日期仅用于表结构兼容。"
        )
    else:
        out.attrs["deadline_column_missing"] = False
    return out


CTDNA_GROUPS = {"LC224", "PC122", "PC228", "PC665"}

# 截图表格"不同样本类型混杂"对应的投入量（ng/例）
# key: 文库类型（从杂交组别/项目前缀推断）
# value: (组织/血/骨髓, CT, 胚系)
MIX_INPUT_NG: dict[str, dict[str, float]] = {
    "122": {"tissue": 800.0, "germline": 200.0},
    "228": {"tissue": 1200.0, "germline": 200.0},
    "HE180": {"tissue": 800.0, "germline": 200.0},
    # L224 或 L218：组织800；骨髓1200（当前输入表无"骨髓/组织"字段，先按 tissue=800）
    "L224_or_L218": {"tissue": 800.0, "germline": 200.0},
    "AML61": {"tissue": 500.0, "germline": 125.0},
    "CNS": {"tissue": 800.0, "germline": 200.0},
    "HRD": {"tissue": 500.0, "germline": 500.0},
    # 665：组织800；CT 1500；胚系200
    "665": {"tissue": 800.0, "ctDNA": 1500.0, "germline": 200.0},
    "THY116": {"tissue": 500.0, "germline": 200.0},
}

# doc：不在表格里的杂交组别，默认按 组织800，胚系200（CT 无默认混杂）
DEFAULT_MIX_INPUT_NG: dict[str, float] = {"tissue": 800.0, "germline": 200.0}


@dataclass(frozen=True)
class Record:
    idx: int  # dataframe 行号
    sample_id: str
    deadline: datetime
    exp_id: str
    conc: float
    grade: str
    vol_ul: float
    hole: str
    index_id: str
    data_g: float
    prefix: str
    hybrid_group: str
    sample_type: str  # tissue/ctDNA/germline
    load_ng: float  # 杂交投入量（ng）
    total_ng: float  # 文库总量（ng），用于"全投/不足则全投"
    deadline_missing_mode: bool = False  # True：无截止日期列，仅用实验编号「#」优先


def _safe_str(x) -> str:
    if x is None or (isinstance(x, float) and pd.isna(x)):
        return ""
    return str(x).strip()


def _sample_has_hash_a_priority(exp_id: str) -> bool:
    return "#" in _safe_str(exp_id)


def _parse_datetime(x) -> datetime:
    if isinstance(x, datetime):
        return x
    # pandas Timestamp
    if hasattr(x, "to_pydatetime"):
        return x.to_pydatetime()
    return pd.to_datetime(x).to_pydatetime()


def _prefix_from_exp_id(exp_id: str) -> str:
    return exp_id.split("-")[0].strip() if exp_id else ""


def _choose_group(raw_group: str | list[str], exp_id: str) -> str:
    # 少量条目为 list（原表同前缀多组别），也有 str 内含换行
    if isinstance(raw_group, list):
        # 目前策略：优先选与 exp_id 字面包含的项，其次取第一项
        for g in raw_group:
            if isinstance(g, str) and g and g in exp_id:
                return g
        return raw_group[0]
    group = raw_group
    if not isinstance(group, str):
        return "UNKNOWN"
    group = group.strip()
    if "\n" in group:
        parts = [p.strip() for p in group.splitlines() if p.strip()]
        # 常见：THY116-D / THY116-R
        if any(tok in exp_id for tok in ["-R", "_R", "RNA", "rna"]) and len(parts) >= 2:
            return parts[-1]
        return parts[0]
    return group


def _group_raw_identity(raw: str | list[str]) -> tuple:
    """比较多条映射是否指向同一杂交组别（支持 value 为 list）"""
    if isinstance(raw, list):
        return ("list", tuple(sorted(str(x) for x in raw)))
    return ("str", str(raw))


def _fuzzy_prefix_lookup(prefix: str) -> str | list[str] | None:
    """
    前缀未在映射表中精确命中时，对映射表的 key 做模糊匹配：
    1) 存在已知键 k 满足 k.startswith(prefix) —— 线下写法较短（如 S_LGT19W 对应表内 S_LGT19WPD）
    2) 存在已知键 k 满足 prefix.startswith(k) 且 len(k)>=5 —— 前缀多了后缀
    取「最长匹配」的一批候选；仅当它们对应的杂交组别取值完全一致时才采纳，否则返回 None 以免误判。
    """
    if not prefix or len(prefix) < 4:
        return None

    keys = list(PREFIX_TO_GROUP.keys())
    cands = [k for k in keys if k.startswith(prefix)]
    if not cands:
        cands = [k for k in keys if prefix.startswith(k) and len(k) >= 5]
        if not cands:
            return None
        max_k_len = max(len(k) for k in cands)
        if len(prefix) - max_k_len > 8:
            return None

    max_len = max(len(k) for k in cands)
    longest = [k for k in cands if len(k) == max_len]
    vals = {_group_raw_identity(PREFIX_TO_GROUP[k]) for k in longest}
    if len(vals) != 1:
        return None
    return PREFIX_TO_GROUP[longest[0]]


def _hybrid_group_from_exp_id(exp_id: str) -> str:
    exp_id = exp_id or ""
    exp_id = exp_id.split("#")[0]  # 剥离 #1/#A 等后缀用于杂交组匹配，优先级判断仍用原始 exp_id
    exp_up = exp_id.upper()
    # 实验编号前缀优先规则（优先于固化映射表）
    if "QW" in exp_up:
        return "QW"
    if exp_up.startswith("HC79"):
        return "HC79"
    if exp_up.startswith("ECS"):
        return "ECS"
    if exp_up.startswith("BRCA"):
        return "PT122"
    if exp_up.startswith("DR_HE"):
        return "TRS" if "_R" in exp_up else "HE180"
    if exp_up.startswith("THY116"):
        return "THY116-R" if exp_up.endswith("-R") else "THY116-D"
    # 665胚系：PC665-B 优先归 PT665，排不下再回退 PC665（见 make_pools）
    if exp_up.startswith("PC665") and exp_up.endswith("-B"):
        return "PT665"
    # HRD：含 HRD 的 X_ 或 PT665 样本，以 -HRD/-HRD-B 结尾才归 HRD，否则回退默认组别
    if (exp_up.startswith("X_") and "HRD" in exp_up) or exp_up.startswith("PT665"):
        if exp_up.endswith("-HRD") or exp_up.endswith("-HRD-B"):
            return "HRD"
        if exp_up.startswith("X_"):
            return "PT228"
        # PT665 开头不含 -HRD，走后续映射表正常处理

    prefix = _prefix_from_exp_id(exp_id)

    raw = PREFIX_TO_GROUP.get(prefix)
    if raw is None:
        raw = _fuzzy_prefix_lookup(prefix)
    if raw is None:
        # BUGFIX: 无法匹配映射表时的兜底规则
        # 若实验编号含有 -N 或 -P，则取第一个 '-' 前字段作为"杂交组别"
        # 例如：PT122-N254 -> PT122
        if "-N" in exp_up or "-P" in exp_up:
            g = _prefix_from_exp_id(exp_id)
            return g if g else "UNKNOWN"
        return "UNKNOWN"
    return _choose_group(raw, exp_id)


def _sample_type(exp_id: str, hybrid_group: str) -> str:
    exp_id = exp_id or ""
    if exp_id.endswith(("-B", "-G", "-GB")):
        return "germline"
    if hybrid_group in CTDNA_GROUPS or hybrid_group.startswith(("PC", "LC")):
        return "ctDNA"
    return "tissue"


def _is_low_priority(exp_id: str, hybrid_group: str) -> bool:
    exp_id = exp_id or ""
    exp_up = exp_id.upper()
    g = (hybrid_group or "").upper()
    # doc: ECS、HC79、QW 优先级低（实验编号任意位置含 QW 即视为 QW 项目）
    return g.startswith("ECS") or g.startswith("HC79") or "QW" in exp_up


def _max_hybrid_per_pool(exp_id: str, hybrid_group: str) -> int:
    """
    BUGFIX: pool 上限按"杂交组别"判定，而不是按实验编号是否 S_/X_。

    经验规则（结合你当前数据与文档表述）：
    - MRD 单杂
    - PT/PC/LC 这类实体瘤/ctDNA组别按 10 杂上限（例如 PT122/PC122/PT228/PC228/PT665/PC665/LC224）
    - 其他默认 12
    """
    exp_id = exp_id or ""
    g = (hybrid_group or "").upper()
    p = _prefix_from_exp_id(exp_id).upper()
    if "MRD" in exp_id.upper() or p.startswith("MRD") or "MRD" in g:
        return 1
    #if g.startswith("HC79") or g.startswith("ECS"):
    #    return 10
    if g == "TRS":
        return 6
    if g == "THY116-R":
        return 6
    if g.startswith(("PT122", "PT228", "PC122", "PC228")):
        return 10
    return 12


def _calc_load_ng(sample_type: str, grade: str, total_ng: float) -> float:
    grade = (grade or "").strip()
    is_lq = grade.upper() == "LQ"
    # 规则：非LQ => 500；组织LQ => 750；胚系混杂按200（此处统一胚系200）
    if sample_type == "germline":
        target = 200.0
    else:
        target = 750.0 if (sample_type == "tissue" and is_lq) else 500.0
    if total_ng <= 0:
        return 0.0
    return float(min(target, total_ng))


def _mix_key_from_record(r: Record) -> str | None:
    g = (r.hybrid_group or "").upper()
    p = (r.prefix or "").upper()
    # 122/228
    if "122" in g:
        return "122"
    if "228" in g:
        return "228"
    # HE180
    if g.startswith("HE180") or p.startswith("HE180"):
        return "HE180"
    # L224 / L218（映射里常见 L218/LC224）
    if g.startswith("LC224") or g.startswith("L218") or p.startswith("L218") or p.startswith("LC224"):
        return "L224_or_L218"
    # AML61
    if p.startswith("AML61") or g.startswith("AML61"):
        return "AML61"
    # CNS（映射里有 CNST/CNST206 等）
    if g.startswith("CNS") or g.startswith("CNST") or p.startswith(("CNS", "CNST")):
        return "CNS"
    # HRD
    if "HRD" in g or "HRD" in (r.exp_id or "").upper():
        return "HRD"
    # 665
    if "665" in g or p.startswith(("PC665", "PT665")):
        return "665"
    # THY116
    if "THY116" in g or p.startswith("THY116"):
        return "THY116"
    return None


def _desired_load_ng(r: Record, *, mixed: bool) -> float:
    """
    mixed=True: 使用"不同样本类型混杂"表格投入量（并按 total_ng 截断全投规则）
    mixed=False: 使用基础规则（500 / LQ组织750 / 胚系200）
    """
    if not mixed:
        return _calc_load_ng(r.sample_type, r.grade, r.total_ng)

    key = _mix_key_from_record(r)
    table = MIX_INPUT_NG.get(key or "", {})
    target = table.get(r.sample_type)
    if target is None:
        target = DEFAULT_MIX_INPUT_NG.get(r.sample_type)
    if target is None:
        # CT 在非665默认不混，回退到基础规则
        return _calc_load_ng(r.sample_type, r.grade, r.total_ng)
    return float(min(target, r.total_ng))


def read_input_records(df: pd.DataFrame) -> list[Record]:
    missing = [c for c in INPUT_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"输入文件缺少列: {missing}")

    deadline_missing = bool(df.attrs.get("deadline_column_missing", False))
    records: list[Record] = []
    for _, row in df.reset_index(drop=False).iterrows():
        exp_id = _safe_str(row["实验编号"])
        if not exp_id:
            continue
        deadline = _parse_datetime(row["报告截止时间"])
        conc = float(row["预文库浓度（ng/μL）"]) if not pd.isna(row["预文库浓度（ng/μL）"]) else 0.0
        vol_ul = float(row["预文库取样体积（μL）"]) if not pd.isna(row["预文库取样体积（μL）"]) else 0.0
        total_ng = conc * vol_ul

        prefix = _prefix_from_exp_id(exp_id)
        hybrid_group = _hybrid_group_from_exp_id(exp_id)
        st = _sample_type(exp_id, hybrid_group)
        load_ng = _calc_load_ng(st, _safe_str(row["预文库等级"]), total_ng)

        records.append(
            Record(
                idx=int(row["index"]),
                sample_id=_safe_str(row["样本编号"]),
                deadline=deadline,
                exp_id=exp_id,
                conc=conc,
                grade=_safe_str(row["预文库等级"]),
                vol_ul=vol_ul,
                hole=_safe_str(row["孔位"]),
                index_id=_safe_str(row["文库index编号"]),
                data_g=float(row["数据量（G）"]) if not pd.isna(row["数据量（G）"]) else 0.0,
                prefix=prefix,
                hybrid_group=hybrid_group,
                sample_type=st,
                load_ng=load_ng,
                total_ng=float(total_ng),
                deadline_missing_mode=deadline_missing,
            )
        )
    return records


def _pool_sort_key(r: Record):
    # doc: 有截止日期→截止时间优先；无截止日期列→实验编号含「#」优先
    is_lq = (r.grade or "").upper() == "LQ"
    hole_prefix = r.hole.split("-")[0] if r.hole else ""
    if r.deadline_missing_mode:
        ha = 0 if _sample_has_hash_a_priority(r.exp_id) else 1
        return (ha, 0 if is_lq else 1, hole_prefix, r.index_id, r.exp_id)
    return (r.deadline, 0 if is_lq else 1, hole_prefix, r.index_id)


@dataclass
class Pool:
    hybrid_group: str
    max_n: int
    records: list[Record]

    @property
    def total_load(self) -> float:
        return float(sum(r.load_ng for r in self.records))

    def total_load_if_mixed(self, extra: Record | None = None) -> float:
        rs = self.records + ([extra] if extra is not None else [])
        return float(sum(_desired_load_ng(x, mixed=True) for x in rs))

    def total_load_if_not_mixed(self, extra: Record | None = None) -> float:
        rs = self.records + ([extra] if extra is not None else [])
        return float(sum(_desired_load_ng(x, mixed=False) for x in rs))

    def is_mixed(self, extra: Record | None = None) -> bool:
        rs = self.records + ([extra] if extra is not None else [])
        types = {x.sample_type for x in rs}
        return len(types) > 1

    @property
    def n(self) -> int:
        return len(self.records)

    @property
    def index_set(self) -> set[str]:
        return {r.index_id for r in self.records if r.index_id}

    def can_add(self, r: Record) -> bool:
        if self.hybrid_group != r.hybrid_group:
            return False
        if r.index_id and r.index_id in self.index_set:
            return False
        if self.n + 1 > self.max_n:
            return False
        # 若加入后变成混杂pool，则按混杂表格投入量重新计算
        mixed = self.is_mixed(extra=r)
        total = self.total_load_if_mixed(extra=r) if mixed else self.total_load_if_not_mixed(extra=r)
        if total > 6000.0:
            return False
        return True

    def add(self, r: Record) -> None:
        # 若加入后 pool 变成混杂，则把 pool 内所有样本的 load_ng 按混杂表格重算
        mixed = self.is_mixed(extra=r)
        new_records = self.records + [r]
        if mixed:
            new_records = [
                Record(**{**x.__dict__, "load_ng": _desired_load_ng(x, mixed=True)}) for x in new_records
            ]
        else:
            new_records = [
                Record(**{**x.__dict__, "load_ng": _desired_load_ng(x, mixed=False)}) for x in new_records
            ]
        self.records = new_records


def make_pools(records: list[Record]) -> list[Pool]:
    pools: list[Pool] = []

    by_group: dict[str, list[Record]] = {}
    for r in records:
        by_group.setdefault(r.hybrid_group, []).append(r)

    for group, rs in by_group.items():
        # 规则：同样本类型尽量同池；ctDNA(除665)不混其他
        rs_sorted = sorted(rs, key=_pool_sort_key)

        for r in rs_sorted:
            placed = False
            max_n = _max_hybrid_per_pool(r.exp_id, group)

            # 先尝试放入"同类型"的已有pool
            for p in pools:
                if p.hybrid_group != group:
                    continue
                # ctDNA(除665类)严格不混
                is_665 = "665" in (group or "")
                if is_665:
                    # 665类：ctDNA/tissue/germline 三者均可混杂，不限胚系数量
                    pass
                elif r.sample_type == "ctDNA":
                    # 非665的ctDNA pool 只收 ctDNA
                    if any(x.sample_type != "ctDNA" for x in p.records):
                        continue
                else:
                    # 组织/胚系允许混杂：优先同类型，其次允许"胚系≤2"混入组织pool
                    if p.records:
                        p_types = {x.sample_type for x in p.records}
                        if r.sample_type in p_types:
                            pass
                        else:
                            # 仅允许 germline <-> tissue 混杂
                            if not ({r.sample_type} | p_types <= {"tissue", "germline"}):
                                continue
                            # germline ≤2（无论是新增还是已有）
                            germ_n = sum(1 for x in p.records if x.sample_type == "germline") + (
                                1 if r.sample_type == "germline" else 0
                            )
                            if germ_n > 2:
                                continue
                if p.can_add(r):
                    p.add(r)
                    placed = True
                    break

            if not placed:
                # PC665胚系回退：优先进PT665，PT665满了才放PC665 pool
                if (r.exp_id.upper().startswith("PC665")
                        and r.sample_type == "germline"
                        and group == "PT665"):
                    for p in pools:
                        if p.hybrid_group != "PC665":
                            continue
                        if r.index_id and r.index_id in p.index_set:
                            continue
                        if p.n + 1 > p.max_n:
                            continue
                        mixed = p.is_mixed(extra=r)
                        total = (p.total_load_if_mixed(extra=r) if mixed
                                 else p.total_load_if_not_mixed(extra=r))
                        if total <= 6000.0:
                            p.add(r)
                            placed = True
                            break
            if not placed:
                pools.append(Pool(hybrid_group=group, max_n=max_n, records=[r]))

        # 665 类允许 ctDNA / tissue 混杂；在上面的逻辑中已放宽 is_665

    # 二次整理：同组内按截止时间或「#A」规则排序，保证相邻行
    def _pool_primary_order(p: Pool):
        rs = p.records
        if not rs:
            return (datetime.max,)
        if rs[0].deadline_missing_mode:
            return min(
                ((0 if _sample_has_hash_a_priority(r.exp_id) else 1), r.exp_id) for r in rs
            )
        return (min(r.deadline for r in rs),)

    pools = sorted(pools, key=lambda p: (p.hybrid_group, _pool_primary_order(p)))
    return pools


def pool_message(p: Pool, now: datetime) -> str:
    # BUGFIX: pool 满杂则提示信息为空
    if p.n >= p.max_n or p.total_load >= 6000:
        return ""
    due_cutoff = (now + timedelta(days=3)).date()
    due_n = sum(1 for r in p.records if hasattr(r.deadline, 'date') and not pd.isna(r.deadline) and r.deadline.date() <= due_cutoff)
    remain_n = max(p.max_n - p.n, 0)
    remain_ng = max(6000.0 - p.total_load, 0.0)
    base = ""
    if due_n > 0:
        # BUGFIX: 括号里的解释不要输出
        base += f"这个pool里有{due_n}个到期样本，但是"
    else:
        base += "这个pool没有到期样本，但是"
    base += f"该pool未满，还可以混入{remain_n}个下一批次样本，或是投入{int(round(remain_ng))}ng文库。"
    return base


def assign_hybrid_ids(pools: list[Pool], now: datetime) -> dict[int, tuple[str, str]]:
    """
    返回: row_idx -> (杂交编号, 提示信息)
    """
    date_tag = now.strftime("%y%m%d")
    counter: dict[str, int] = {}
    out: dict[int, tuple[str, str]] = {}
    for p in pools:
        counter[p.hybrid_group] = counter.get(p.hybrid_group, 0) + 1
        n = counter[p.hybrid_group]
        g = p.hybrid_group
        if g == "THY116-R":
            hid = f"THY116-{date_tag}-R{n}"
        elif g == "THY116-D":
            hid = f"THY116-{date_tag}-D{n}"
        else:
            hid = f"{g}-{date_tag}-{n}"
        msg = pool_message(p, now)
        for r in p.records:
            out[r.idx] = (hid, msg)
    return out


def _has_duplicate_indexes(records: Iterable[Record]) -> bool:
    seen: set[str] = set()
    for r in records:
        if not r.index_id:
            continue
        if r.index_id in seen:
            return True
        seen.add(r.index_id)
    return False


def assign_lines_for_pools(
    records: list[Record],
    hybrid_info: dict[int, tuple[str, str]],
) -> dict[int, tuple[str, str]]:
    """
    按「杂交 pool」（同一杂交编号）整体分配 line：同一 pool 内所有样本 line 一致。
    返回: row_idx -> (line号, extra_msg_append)；排不下时 line 为空串，extra 含「该批次排不下」。
    """
    dup = _has_duplicate_indexes(records)

    by_hid: dict[str, list[Record]] = defaultdict(list)
    for r in records:
        hid, _ = hybrid_info.get(r.idx, ("", ""))
        key = hid if hid else f"__orphan_{r.idx}"
        by_hid[key].append(r)

    pool_items = list(by_hid.items())

    def _pool_sort_key(item: tuple[str, list[Record]]):
        hid, rs = item
        if rs and rs[0].deadline_missing_mode:
            r0 = min(
                rs,
                key=lambda x: (
                    _is_low_priority(x.exp_id, x.hybrid_group),
                    0 if _sample_has_hash_a_priority(x.exp_id) else 1,
                    x.exp_id,
                ),
            )
            return (
                _is_low_priority(r0.exp_id, r0.hybrid_group),
                0 if _sample_has_hash_a_priority(r0.exp_id) else 1,
                hid,
            )
        r0 = min(rs, key=lambda x: (x.deadline, x.exp_id))
        return (_is_low_priority(r0.exp_id, r0.hybrid_group), r0.deadline, hid)

    pool_items.sort(key=_pool_sort_key)

    out: dict[int, tuple[str, str]] = {}

    if not dup:
        total = 0.0
        for hid, rs in pool_items:
            vol = sum(r.data_g for r in rs)
            if total + vol <= 850.0:
                total += vol
                for r in rs:
                    out[r.idx] = ("line1", "")
            else:
                for r in rs:
                    out[r.idx] = ("", "该批次排不下")
        return out

    line_caps = [200.0, 200.0, 200.0, 200.0]
    line_vols = [0.0, 0.0, 0.0, 0.0]
    line_indexes: list[set[str]] = [set(), set(), set(), set()]

    for hid, rs in pool_items:
        vol = sum(r.data_g for r in rs)
        idx_vals = [r.index_id for r in rs if r.index_id]
        if len(idx_vals) != len(set(idx_vals)):
            for r in rs:
                out[r.idx] = ("", "该批次排不下")
            continue
        idx_set = set(idx_vals)

        candidates: list[int] = []
        for li in range(4):
            if line_vols[li] + vol > line_caps[li]:
                continue
            if idx_set & line_indexes[li]:
                continue
            candidates.append(li)
        if not candidates:
            for r in rs:
                out[r.idx] = ("", "该批次排不下")
            continue
        best = min(candidates, key=lambda li: line_vols[li])
        line_vols[best] += vol
        line_indexes[best] |= idx_set
        for r in rs:
            out[r.idx] = (f"line{best+1}", "")
    return out


def merge_hybrid_cells(xlsx_path: Path, sheet: str = "Sheet1") -> None:
    wb = load_workbook(xlsx_path)
    ws = wb[sheet]

    # 找"杂交编号"列
    header_row = 1
    col_idx = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(row=header_row, column=c).value == "杂交编号":
            col_idx = c
            break
    if col_idx is None:
        wb.save(xlsx_path)
        return

    start = 2
    last_val = ws.cell(row=2, column=col_idx).value
    for r in range(3, ws.max_row + 2):
        val = ws.cell(row=r, column=col_idx).value if r <= ws.max_row else None
        if val != last_val:
            if last_val not in (None, "") and r - 1 > start:
                ws.merge_cells(start_row=start, start_column=col_idx, end_row=r - 1, end_column=col_idx)
            start = r
            last_val = val

    wb.save(xlsx_path)


def main(argv: list[str] | None = None) -> None:
    base = Path(__file__).resolve().parent
    args = argv if argv is not None else sys.argv[1:]
    if args:
        in_path = Path(args[0]).expanduser()
        if not in_path.is_absolute():
            in_path = base / in_path
    else:
        in_path = base / "输入文件.xlsx"
    if not in_path.exists():
        raise FileNotFoundError(f"找不到输入文件: {in_path}")

    if len(args) > 1:
        out_path = Path(args[1]).expanduser()
        if not out_path.is_absolute():
            out_path = base / out_path
    else:
        out_path = base / "输出文件.xlsx"

    xf = pd.ExcelFile(in_path)
    sheet = "Sheet1" if "Sheet1" in xf.sheet_names else xf.sheet_names[0]
    df = pd.read_excel(in_path, sheet_name=sheet)
    df = normalize_input_dataframe(df)
    # 保留原行序，但我们会在输出阶段把同一杂交编号尽量放一起
    records = read_input_records(df)
    now = datetime.now()

    pools = make_pools(records)
    hybrid_info = assign_hybrid_ids(pools, now)  # idx -> (hid, msg)
    line_info = assign_lines_for_pools(records, hybrid_info)  # idx -> (line, extra_msg)

    # 写回到df
    df_out = df.copy()
    df_out["杂交编号"] = ""
    df_out["提示信息"] = ""
    df_out["line号"] = ""

    for r in records:
        hid, msg = hybrid_info.get(r.idx, ("", ""))
        line, extra = line_info.get(r.idx, ("line1", ""))
        df_out.at[r.idx, "杂交编号"] = hid
        if extra:
            tip = (msg + "；" + extra) if msg else extra
        else:
            tip = msg
        df_out.at[r.idx, "提示信息"] = tip
        if "该批次排不下" in _safe_str(tip):
            df_out.at[r.idx, "line号"] = ""
        else:
            df_out.at[r.idx, "line号"] = line if line else "line1"

    # 排序：line 号有内容的行在前；其次同杂交编号相邻
    def _line_nonempty_rank(v) -> int:
        if v is None or (isinstance(v, float) and pd.isna(v)):
            return 1
        return 0 if str(v).strip() else 1

    df_out["_line_nonempty_first"] = df_out["line号"].map(_line_nonempty_rank)
    deadline_col_missing = bool(df.attrs.get("deadline_column_missing", False))
    if deadline_col_missing:
        df_out["_sort_pool_order"] = df_out["实验编号"].map(
            lambda x: 0 if _sample_has_hash_a_priority(_safe_str(x)) else 1
        )
    else:
        df_out["_sort_pool_order"] = pd.to_datetime(df_out["报告截止时间"], errors="coerce")
    df_out = df_out.sort_values(
        by=["_line_nonempty_first", "杂交编号", "_sort_pool_order"],
        kind="mergesort",
    )
    df_out = df_out.drop(columns=["_line_nonempty_first", "_sort_pool_order"])

    # 保证列顺序与模板一致（多余列保留在后面）
    for col in OUTPUT_COLUMNS:
        if col not in df_out.columns:
            df_out[col] = ""
    ordered = OUTPUT_COLUMNS + [c for c in df_out.columns if c not in OUTPUT_COLUMNS]
    df_out = df_out[ordered]

    try:
        df_out.to_excel(out_path, index=False, sheet_name="Sheet1")
        merge_hybrid_cells(out_path, sheet="Sheet1")
    except PermissionError:
        # 常见原因：Excel 正在打开输出文件导致无法覆盖写入
        fallback = base / f"输出文件_生成_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        df_out.to_excel(fallback, index=False, sheet_name="Sheet1")
        merge_hybrid_cells(fallback, sheet="Sheet1")
        print(f"[WARN] 无法覆盖写入 {out_path.name}（可能被占用），已输出到：{fallback.name}")


if __name__ == "__main__":
    main()

